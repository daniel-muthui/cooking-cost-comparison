"""
clean_cct_data.py
Cleans the raw eCAP Controlled Cooking Test workbook (Data_from_eCap_CCTs.xlsx)
into a validated lookup table (cct_lookup.csv) for the Household Energy and
Cost Comparison Tool.

What it does, in order:
 1. Reads Sheet1 (45 test records).
 2. Normalizes appliance and dish names to a controlled vocabulary.
 3. Parses per-test energy readings: kWh for electric appliances, grams of
    fuel for non-electric.
 4. Excludes documented outlier readings (listed in EXCLUSIONS below).
 5. Recomputes kWh-equivalent from raw fuel mass and calorific values --
    the stored "KWh Equivalent" column contains transcription errors
    (e.g. ethanol-rice 6.117 kWh) and is NOT trusted.
 6. Recomputes cooking time per test from start/end timestamps and averages
    valid tests -- the stored time columns are internally inconsistent.
 7. Emits cct_lookup.csv with one row per appliance-dish combination and a
    notes column recording every correction or exclusion applied.

Every number in the output traces back to a raw cell; this script is the
data-provenance record. Run: python clean_cct_data.py <input.xlsx>
"""

import csv
import re
import sys
from datetime import time as dtime
from openpyxl import load_workbook

# ---------------------------------------------------------------- constants

# Calorific values (MJ/kg) -- engineeringtoolbox.com, as used in the eCAP report
CALORIFIC_MJ_PER_KG = {
    "charcoal": 31.0,
    "lpg": 46.1,
    "ethanol": 27.0,
    "kerosene": 43.1,
}
MJ_TO_KWH = 1 / 3.6  # exact; the report's 0.2778 is this rounded

# Controlled vocabulary: raw name fragments -> (canonical appliance, fuel)
APPLIANCE_MAP = [
    ("epc", ("EPC", "electricity")),
    ("induction", ("Induction cooker", "electricity")),
    ("rice cooker", ("Rice cooker", "electricity")),
    ("air", ("Air fryer", "electricity")),      # catches "Airfyer"
    ("infrared", ("Infrared cooker", "electricity")),
    ("hot plate", ("Hot plate", "electricity")),
    ("lpg", ("LPG stove", "lpg")),
    ("charcoal", ("Improved charcoal stove (ICS)", "charcoal")),
    ("ethanol", ("Ethanol stove", "ethanol")),
    ("kerosene", ("Kerosene stove", "kerosene")),
]

DISH_MAP = {
    "rice": "Rice",
    "beans": "Beans",
    "spinach": "Spinach",
    "beef": "Beef stew",
    "ugali": "Ugali",
    "chapati": "Chapati",
    "chips": "Chips",
}

# Documented exclusions: (dish, appliance, test_number) -> reason.
# Ethanol-chips test 1 recorded 1015 g against 280 g / 220 g in tests 2-3;
# the study protocol (third test when first two diverge >15%) implies the
# third test replaced it.
EXCLUSIONS = {
    ("Chips", "Ethanol stove", 1): "1015 g outlier vs 280/220 g in tests 2-3; excluded per study's own divergence protocol",
}

# Values kept as recorded but flagged for review before relying on them.
REVIEW_FLAGS = {
    ("Chapati", "LPG stove"): "161 g LPG recorded vs 20-30 g on other fuels; likely transcription error in raw data (appears in eCAP report too) -- REVIEW",
}

DATA_START_ROW = 10  # 1-indexed; rows above are legend + headers


# ------------------------------------------------------------------ parsing

def canonical_appliance(raw):
    s = re.sub(r"\s+", " ", str(raw)).strip().lower()
    for frag, result in APPLIANCE_MAP:
        if frag in s:
            return result
    raise ValueError(f"Unrecognized appliance name: {raw!r}")


def canonical_dish(raw):
    s = str(raw).strip().lower()
    for frag, name in DISH_MAP.items():
        if frag in s:
            return name
    raise ValueError(f"Unrecognized dish name: {raw!r}")


def parse_weight_g(raw):
    """'364g' -> 364.0; '-', '', None, whitespace -> None."""
    if raw is None:
        return None
    s = str(raw).strip().lower().replace("g", "").replace(",", "")
    if s in ("", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_energy(raw, fuel):
    """Return (value, unit) where unit is 'kwh' or 'g', or None if absent.

    Electric rows store '0.278kwh' (sometimes bare numbers); fuel rows store
    '35g' (sometimes bare numbers = grams, e.g. kerosene-spinach '25.0').
    """
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if s in ("", "-"):
        return None
    has_kwh = "kwh" in s
    has_g = s.endswith("g")
    num = re.sub(r"[^0-9.]", "", s)
    if not num:
        return None
    val = float(num)
    if has_kwh or (fuel == "electricity" and not has_g):
        return (val, "kwh")
    return (val, "g")


def to_day_fraction(v):
    """Timestamps arrive as datetime.time; convert to fraction of a day."""
    if isinstance(v, dtime):
        return (v.hour * 3600 + v.minute * 60 + v.second) / 86400.0
    if isinstance(v, (int, float)):
        return float(v)
    return None


def test_duration_min(start, end):
    """Duration in minutes from two time-of-day values; None if unusable.
    Assumes no test crossed midnight (all recorded times are daytime)."""
    s, e = to_day_fraction(start), to_day_fraction(end)
    if s is None or e is None:
        return None
    d = (e - s) * 24 * 60
    return d if d > 0 else None


# -------------------------------------------------------------------- main

def clean(path_in, path_out):
    wb = load_workbook(path_in, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    records = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        food_raw, cooker_raw = row[2], row[3]
        if not food_raw or not cooker_raw:
            continue  # blank/legend rows

        appliance, fuel = canonical_appliance(cooker_raw)
        dish = canonical_dish(food_raw)
        notes = []

        # --- energy: parse tests 1-3, apply exclusions, average -----------
        energies_kwh, fuel_g_list = [], []
        for t, cell in enumerate(row[5:8], start=1):
            parsed = parse_energy(cell, fuel)
            if parsed is None:
                continue
            reason = EXCLUSIONS.get((dish, appliance, t))
            if reason:
                notes.append(f"test {t} excluded: {reason}")
                continue
            val, unit = parsed
            if unit == "kwh":
                energies_kwh.append(val)
            else:
                fuel_g_list.append(val)

        n_tests = len(energies_kwh) + len(fuel_g_list)
        if fuel == "electricity":
            energy_kwh = round(sum(energies_kwh) / len(energies_kwh), 4) if energies_kwh else None
            fuel_mass_g = None
        else:
            if fuel_g_list:
                fuel_mass_g = round(sum(fuel_g_list) / len(fuel_g_list), 1)
                energy_kwh = round(
                    (fuel_mass_g / 1000.0) * CALORIFIC_MJ_PER_KG[fuel] * MJ_TO_KWH, 4
                )
                notes.append("kWh recomputed from fuel mass x calorific value; stored 'KWh Equivalent' column not used")
            else:
                fuel_mass_g, energy_kwh = None, None

        # flag where the stored kWh disagrees materially with the recomputed one
        stored_kwh = row[10]
        if (
            isinstance(stored_kwh, (int, float)) and energy_kwh
            and abs(stored_kwh - energy_kwh) / energy_kwh > 0.10
        ):
            notes.append(f"stored kWh ({stored_kwh}) differed >10% from recomputed ({energy_kwh})")

        # --- time: recompute from timestamps ------------------------------
        durations = []
        for t, (s_col, e_col) in enumerate([(11, 12), (13, 14), (15, 16)], start=1):
            if EXCLUSIONS.get((dish, appliance, t)):
                continue
            d = test_duration_min(row[s_col], row[e_col])
            if d is not None:
                durations.append(d)
        time_min = round(sum(durations) / len(durations), 1) if durations else None
        if durations and len(durations) != n_tests:
            notes.append(f"time averaged over {len(durations)} tests vs {n_tests} energy tests (missing/invalid timestamps)")
        if not durations:
            notes.append("no usable start/end timestamps; time missing")

        review = REVIEW_FLAGS.get((dish, appliance))
        if review:
            notes.append(review)

        # cross-test divergence flag (study's own >15% criterion)
        vals = energies_kwh or fuel_g_list
        if len(vals) >= 2 and min(vals) > 0 and (max(vals) - min(vals)) / min(vals) > 0.15:
            notes.append("energy tests diverge >15%; interpret with caution")

        records.append({
            "dish": dish,
            "appliance": appliance,
            "fuel": fuel,
            "food_weight_g": parse_weight_g(row[4]),
            "n_energy_tests": n_tests,
            "fuel_mass_g": fuel_mass_g,
            "energy_kwh": energy_kwh,
            "time_min": time_min,
            "ease_of_use": row[19],
            "taste": row[20],
            "notes": "; ".join(notes),
        })

    # stable ordering: dish, then appliance
    records.sort(key=lambda r: (r["dish"], r["appliance"]))

    with open(path_out, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(records[0].keys()))
        writer.writeheader()
        writer.writerows(records)

    return records


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "Data_from_eCap_CCTs.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "cct_lookup.csv"
    recs = clean(src, out)
    print(f"Wrote {len(recs)} appliance-dish records to {out}")
    missing = [r for r in recs if r["energy_kwh"] is None or r["time_min"] is None]
    if missing:
        print("Records with missing energy or time:")
        for r in missing:
            print(f"  {r['dish']} / {r['appliance']}: {r['notes']}")
